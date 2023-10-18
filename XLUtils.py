import openpyxl

def getRowCount(file, sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetName)
    return (sheet.max_row)

def getColunmCount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    return (sheet.max_column)

def redData(file,sheetname,rownum,columnnum):
    workboom=openpyxl.load_workbook(file)
    sheet=workboom.get_sheet_by_name(sheetname)
    return sheet.cell(row=rownum, column=columnnum).value

def writedata(file,sheetname,rownum,columnnum, data):
    workboom=openpyxl.load_workbook(file)
    sheet=workboom.get_sheet_by_name(sheetname)
    sheet.cell(row=rownum, column=columnnum).value=data
    workboom.save(file)