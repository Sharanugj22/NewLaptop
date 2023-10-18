import pickle
import random

import openpyxl
#Read data from Excel sheet

# path="C:\\Users\\DELL\Desktop\\Test.xlsx"
# wb=openpyxl.load_workbook(path)
# sheet=wb.active
# # title=wb.get_sheet_by_name("Sheet1")
# rows=sheet.max_row
# cols=sheet.max_column
# print(("Max row counr:",rows))
# print("Max column count:",cols)
# for r in range(1,rows+1):
#     for c in range(1,cols + 1):
#         dd=sheet.cell(row=r, column=c)
#         print(dd.value, end=" ")
#     print()

# Write date into Excel sheet

# path="C:\\Users\\DELL\Desktop\\Test.xlsx"
# workbook=openpyxl.load_workbook(path)
# sheet=workbook.active
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(row=r, column=c).value="welcome"
# workbook.save(path)
num=[1,2,3]

num=num+[4,5,6]
ss=num
print(ss, "dd", num)

